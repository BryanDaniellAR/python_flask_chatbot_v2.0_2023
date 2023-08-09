import openpyxl
import pandas as pd

def tienda_data(path_file,hoja):
    try:
        data = pd.read_excel(path_file,sheet_name=hoja)
        df_data = pd.DataFrame(data)
        return df_data
    except:
        return [{"error":"Sucedio un error en la tienda de información"}]

def new_reserva (path_file,hoja):
    # Cargar el archivo Excel completo
    xls = pd.ExcelFile(path_file)
    hojas = {}
    for sheet_name in xls.sheet_names:
        hojas[sheet_name] = pd.read_excel(xls, sheet_name)
    df = hojas[hoja]
    ultimo_id = df['id'].max() if not df.empty else 0
    nuevo_id = ultimo_id + 1
    nueva_fila = {'id': nuevo_id, 'nombre': '', 'numero': '', 'email': '', 'producto': '', 'cantidad': 0} 
    df = df.append(nueva_fila, ignore_index=True)
    hojas[hoja] = df
    with pd.ExcelWriter(path_file) as writer:
        for sheet_name, sheet_df in hojas.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

def new_reserva_nombre (path_file,hoja,msg):
    xls = pd.ExcelFile(path_file)
    hojas = {}
    for sheet_name in xls.sheet_names:
        hojas[sheet_name] = pd.read_excel(xls, sheet_name)
    df = hojas[hoja]
    
    # Modificar el valor del campo "numero" en el último registro
    indice_ultimo_registro = df.index[-1]
    df.at[indice_ultimo_registro, 'nombre'] = msg
    
    # Guardar todas las hojas en un nuevo archivo Excel
    with pd.ExcelWriter(path_file) as writer:
        for sheet_name, sheet_df in hojas.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
def new_reserva_numero (path_file,hoja,msg):
    xls = pd.ExcelFile(path_file)
    hojas = {}
    for sheet_name in xls.sheet_names:
        hojas[sheet_name] = pd.read_excel(xls, sheet_name)
    df = hojas[hoja]
    
    # Modificar el valor del campo "numero" en el último registro
    indice_ultimo_registro = df.index[-1]
    df.at[indice_ultimo_registro, 'numero'] = msg
    
    # Guardar todas las hojas en un nuevo archivo Excel
    with pd.ExcelWriter(path_file) as writer:
        for sheet_name, sheet_df in hojas.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
def new_reserva_email (path_file,hoja,msg):
    xls = pd.ExcelFile(path_file)
    hojas = {}
    for sheet_name in xls.sheet_names:
        hojas[sheet_name] = pd.read_excel(xls, sheet_name)
    df = hojas[hoja]
    
    # Modificar el valor del campo "numero" en el último registro
    indice_ultimo_registro = df.index[-1]
    df.at[indice_ultimo_registro, 'email'] = msg
    
    # Guardar todas las hojas en un nuevo archivo Excel
    with pd.ExcelWriter(path_file) as writer:
        for sheet_name, sheet_df in hojas.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
def new_reserva_producto (path_file,hoja,msg):
    xls = pd.ExcelFile(path_file)
    hojas = {}
    for sheet_name in xls.sheet_names:
        hojas[sheet_name] = pd.read_excel(xls, sheet_name)
    df = hojas[hoja]
    
    # Modificar el valor del campo "numero" en el último registro
    indice_ultimo_registro = df.index[-1]
    df.at[indice_ultimo_registro, 'producto'] = msg
    
    # Guardar todas las hojas en un nuevo archivo Excel
    with pd.ExcelWriter(path_file) as writer:
        for sheet_name, sheet_df in hojas.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
def new_reserva_cantidad (path_file,hoja,msg):
    xls = pd.ExcelFile(path_file)
    hojas = {}
    for sheet_name in xls.sheet_names:
        hojas[sheet_name] = pd.read_excel(xls, sheet_name)
    df = hojas[hoja]
    
    # Modificar el valor del campo "numero" en el último registro
    indice_ultimo_registro = df.index[-1]
    df.at[indice_ultimo_registro, 'cantidad'] = msg
    
    # Guardar todas las hojas en un nuevo archivo Excel
    with pd.ExcelWriter(path_file) as writer:
        for sheet_name, sheet_df in hojas.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

def agregar_reserva(path_file,hoja):
    Reservacion = pd.read_excel(path_file,sheet_name=hoja)
    df_Reservacion = pd.DataFrame(Reservacion)
    list_Reservacion_id = df_Reservacion.id.to_list()
    reserva_contador_id = len(list_Reservacion_id)
    wbk = openpyxl.load_workbook(path_file)
    wks = wbk[hoja]
    wks.append([reserva_contador_id+1])
    wbk.save(path_file)
    wbk.close

def modificar_reserva(path_file,hoja,informacion_recibida):
    Reservacion = pd.read_excel(path_file,sheet_name=hoja)
    df_Reservacion = pd.DataFrame(Reservacion)
    list_Reservacion_id = df_Reservacion.id.to_list()
    list_Reservacion_nombres = df_Reservacion.nombre.to_list()
    list_Reservacion_numero = df_Reservacion.numero.to_list()
    list_Reservacion_email = df_Reservacion.email.to_list()
    list_Reservacion_producto = df_Reservacion.producto.to_list()
    list_Reservacion_cantidad = df_Reservacion.cantidad.to_list()
    reserva_contador_id = len(list_Reservacion_id)
    wbk = openpyxl.load_workbook(path_file)
    wks = wbk[hoja]

    if pd.isna(list_Reservacion_nombres[len(list_Reservacion_nombres)-1]):
        wks.cell(row=reserva_contador_id+1, column=2).value = informacion_recibida
        wbk.save(path_file)
        wbk.close
        return True
    elif pd.isna(list_Reservacion_numero[len(list_Reservacion_numero)-1]):
        wks.cell(row=reserva_contador_id+1, column=3).value = informacion_recibida
        wbk.save(path_file)
        wbk.close
        return True
    elif pd.isna(list_Reservacion_email[len(list_Reservacion_email)-1]):
        wks.cell(row=reserva_contador_id+1, column=4).value = informacion_recibida
        wbk.save(path_file)
        wbk.close
        return True
    elif pd.isna(list_Reservacion_producto[len(list_Reservacion_producto)-1]):
        wks.cell(row=reserva_contador_id+1, column=5).value = informacion_recibida
        wbk.save(path_file)
        wbk.close
        return True
    elif pd.isna(list_Reservacion_cantidad[len(list_Reservacion_cantidad)-1]):
        wks.cell(row=reserva_contador_id+1, column=6).value = informacion_recibida
        wbk.save(path_file)
        wbk.close
        return True
    else:
        return False